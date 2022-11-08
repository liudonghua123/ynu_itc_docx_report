#%%
from os.path import dirname, join, realpath, basename, sep, exists
import tempfile
from docxtpl import DocxTemplate, InlineImage

# for height and width you have to use millimeters (Mm), inches or points(Pt) class :
from docx.shared import Mm, Cm
import fire
import jinja2
import pandas as pd
from dataclasses import dataclass
from datetime import datetime
from openpyxl import load_workbook
from urllib.parse import urlparse
import requests

#%%
from config_logging import init_logging

logger = init_logging(join(dirname(realpath(__file__)), "main.log"))

#%%
# define a class to hold the data of each row
@dataclass
class Record:
    name: str
    gender: str
    id_num: str
    telphone: str
    company: str
    car_num: str
    access_location: str
    access_date: str
    access_duration: str
    reason: str
    health_code_image: InlineImage | None
    travel_card_image: InlineImage | None
    nucleic_acid_testing_image: InlineImage | None
    health_pledge_image: str | None
    health_status: str = "健康"


#%%
# patch for pandas.read_excel to support hyperlink
# see https://github.com/pandas-dev/pandas/issues/13439#issuecomment-1025641177
from pandas.io.excel._openpyxl import OpenpyxlReader
import numpy as np
from pandas._typing import (
    FilePath,
    ReadBuffer,
    Scalar,
)


def _convert_cell(self, cell, convert_float: bool) -> Scalar:
    from openpyxl.cell.cell import (
        TYPE_ERROR,
        TYPE_NUMERIC,
    )

    # here we adding this hyperlink support:
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
        # just for example, you able to return both value and hyperlink,
        # comment return above and uncomment return below
        # btw this may hurt you on parsing values, if symbols "|||" in value or hyperlink.
        # return f'{cell.value}|||{cell.hyperlink.target}'
    # here starts original code, except for "if" became "elif"
    elif cell.value is None:
        return ""  # compat with xlrd
    elif cell.data_type == TYPE_ERROR:
        return np.nan
    elif not convert_float and cell.data_type == TYPE_NUMERIC:
        return float(cell.value)
    return cell.value


def load_workbook_patch(self, filepath_or_buffer: FilePath | ReadBuffer[bytes]):
    from openpyxl import load_workbook

    # had to change read_only to False (note: may affect on speed of reading file):
    return load_workbook(filepath_or_buffer, read_only=False, data_only=True, keep_links=False)


OpenpyxlReader._convert_cell = _convert_cell
OpenpyxlReader.load_workbook = load_workbook_patch


#%%
# utilities
def get_hyperlink(cell):
    result = None
    try:
        if cell != None and cell.hyperlink != None:
            result = cell.hyperlink.target
    except:
        result = None
    return result


def image_url_to_inline_image(tpl, image_url, width=60, height=None):
    logger.info(f"image_url_to_inline_image: {image_url}")
    inline_image = None
    # image_url == np.nan is always False, use isinstance to check here.
    if image_url == None or not isinstance(image_url, str):
        return None
    url_parse_result = urlparse(image_url)
    suffix = url_parse_result.path.split(".")[-1]
    saved_image_path = f"images/{basename(url_parse_result.path)}"
    if not exists(saved_image_path):
        logger.info(f"Try to fetch image url: {image_url}, suffix: {suffix}")
        response = requests.get(image_url)
        if response.status_code == 200:
            # save the image to local tempfile
            # tmp_img_file = tempfile.NamedTemporaryFile(delete=None, suffix=f".{suffix}")
            # tmp_img_file.write(response.content)
            # tmp_img_file.close()
            # logger.info(f"create tmp image {tmp_img_file.name.replace(sep, '/')} for {image_url}")
            with open(saved_image_path, "wb") as f:
                f.write(response.content)
            logger.info(f"save image {saved_image_path} for {image_url}")
        else:
            logger.error(f"Failed to fetch image url: {image_url}, status code: {response.status_code}")
            saved_image_path = None
    else:
        logger.info(f"image url {image_url} has cached image {saved_image_path}, use it directly")
    inline_image = InlineImage(
        tpl,
        saved_image_path,
        width=Mm(width) if width != None else None,
        height=Mm(height) if height != None else None,
    )
    return inline_image


#%%
def read_excel_by_pandas(input_file_path):
    # read the input data from the excel file using pandas.read_excel
    df = pd.read_excel(input_file_path)
    logger.info(f"Read {len(df)} records from the excel file")
    logger.info(f"Head of the records:\n{df.head()}")
    # iterate the rows of the dataframe
    records: list[Record] = []
    for index, row in df.iterrows():
        # create a Record object
        record = Record(
            name=row["姓名（必填）"],
            gender=row["性别（必填）"],
            id_num=row["身份证号码（必填）"],
            telphone=row["手机号码（必填）"],
            company=row["单位名称（必填）"],
            car_num=row["车牌号码"],
            access_location=row["到访地点（必填）"],
            access_date=row["到访日期（必填）"],
            access_duration=row["入校期限（必填）"],
            reason=row["到访原因（必填）"],
            health_code_image=row["云南省健康码（必填）"],
            travel_card_image=row["行程卡截图（必填）"],
            nucleic_acid_testing_image=row["核酸检测截图（必填）"],
            health_pledge_image=row["《个人健康承诺书》（必填）"],
        )
        records.append(record)
    logger.info(f"Read {len(records)} records from the excel file.")
    return records


def read_excel_by_openpyxl(input_file_path):
    # read the input data from the excel file using openpyxl
    workbook = load_workbook(input_file_path)
    sheet = workbook.active
    # iterate the rows of the worksheet
    records: list[Record] = []
    for (
        _,
        _,
        name,
        gender,
        id_num,
        telphone,
        company,
        car_num,
        access_location,
        access_date,
        access_duration,
        reason,
        health_code_image,
        _,
        travel_card_image,
        _,
        nucleic_acid_testing_image,
        health_pledge_image,
    ) in sheet.iter_rows(min_row=2, max_col=18):
        # create a Record object
        record = Record(
            name=name.value,
            gender=gender.value,
            id_num=id_num.value,
            telphone=telphone.value,
            company=company.value,
            car_num=car_num.value,
            access_location=access_location.value,
            access_date=access_date.value,
            access_duration=access_duration.value,
            reason=reason.value,
            health_code_image=get_hyperlink(health_code_image),
            travel_card_image=get_hyperlink(travel_card_image),
            nucleic_acid_testing_image=get_hyperlink(nucleic_acid_testing_image),
            health_pledge_image=get_hyperlink(health_pledge_image),
        )
        records.append(record)
    logger.info(f"Read {len(records)} records from the excel file.")
    return records


#%%
def main(
    input_file_path: str = join(dirname(realpath(__file__)), "sample.xlsx"),
    output_file_path: str = join(dirname(realpath(__file__)), "generated_doc.docx"),
):

    # generate the output docx file from the template with the data
    tpl = DocxTemplate("template.docx")
    
    # read the input data from the excel file
    records = read_excel_by_pandas(input_file_path)
    # records = read_excel_by_openpyxl(input_file_path)
    
    # update the image url to inline image
    for record in records:
        record.health_code_image = image_url_to_inline_image(tpl, record.health_code_image)
        record.travel_card_image = image_url_to_inline_image(tpl, record.travel_card_image)
        record.nucleic_acid_testing_image = image_url_to_inline_image(tpl, record.nucleic_acid_testing_image)
        record.health_pledge_image = image_url_to_inline_image(tpl, record.health_pledge_image)
    
    # generate the output docx file from the template with the data
    context = {"records": records}
    # pass some global utility functions/objects to the template
    jinja_env = jinja2.Environment(autoescape=True)
    jinja_env.globals["len"] = len
    jinja_env.globals["datetime"] = datetime
    jinja_env.globals["enumerate"] = enumerate
    
    # render the template
    tpl.render(context, jinja_env=jinja_env)
    # save the output docx file
    tpl.save(output_file_path)
    logger.info(f"Save the output docx file to {output_file_path}")


if __name__ == "__main__":
    # Make Python Fire not use a pager when it prints a help text
    fire.core.Display = lambda lines, out: print(*lines, file=out)
    fire.Fire(main)
